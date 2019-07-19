# -*-coding:utf-8 -*-
import requests, time
import re, traceback
from lxml import etree
from openpyxl import Workbook
def tagSearch(offset):
    wb = Workbook()
    ws = wb.active
    ws.append(["number", "link", "description", "address", "external", "verified"])
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['F'].width = 7
    flag = 1
    page = 1
    row = 2
    url = '?filter=16&offset=%d' % ((offset - 1) * 50)
    while flag:
        try:
            surl = 'https://www.blockchain.com/btc/tags' + url
            context = requests.get(surl)
            selector = etree.HTML(context.content, parser=etree.HTMLParser(encoding='utf-8'))
            url = selector.xpath('//*[@class="next "]/a/@href')
            print("page:%d" % page)
            if len(url) > 0:
                url = str(url[0])
                page += 1
            else:
                flag = 0
            entry_list = selector.xpath('/html/body/div/table/tbody/tr')
            for entry in entry_list:
                address = entry.xpath('./td[1]/a/text()')
                tag = entry.xpath('./td[2]/span/text()')
                link = entry.xpath('./td[3]/a/text()')
                verify = entry.xpath('./td[4]/img/@src')
                ws.cell(row, 1).value = str(row - 1)
                ws.cell(row, 2).value = str(link).replace('[', '').replace(']', '').replace('\'', '')
                ws.cell(row, 3).value = str(tag).replace('[', '').replace(']', '').replace('\'', '')
                ws.cell(row, 4).value = str(address).replace('[', '').replace(']', '').replace('\'', '')
                description = re.compile(r'(?=red)').findall(str(verify))
                if len(description) == 1:
                    ws.cell(row, 6).value = 'NO'
                else:
                    ws.cell(row, 6).value = 'YES'
                # 含blockchain
                blockchain = re.compile(r'(?=www.blockchain.com)').findall(str(link))
                # 不含blockchain
                if len(blockchain) < 1:
                    ws.cell(row, 5).value = 'Y'
                else:
                    ws.cell(row, 5).value = 'N'
                row += 1
        except Exception as e:
            print("url failed:", e)
            traceback.print_exc()
            continue
    wb.save("messages.xlsx")


# 主函数


if __name__ == '__main__':
    print("start--------------")
    page = 1
    tagSearch(page)
    print("end----------------")


