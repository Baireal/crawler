# -*-coding:utf-8 -*-
#js 反爬，出现521错误。重点在于获取新的cookie。
# 首次访问-获取js代码-执行代码，修改函数使得执行结果可得（var result）-结果仍为一段js代码，执行得到__jsl_clearance。
#拼接__jsluid_s，__jsl_clearance成为cookies，再次访问
#xpath空问题：1.浏览器优化tbody等，需打印源码进行路径选取 2.response.text 和response.content的区别
from lxml import etree
import re,time,traceback              #实现正则表达式
import execjs              #执行js代码
import requests            #爬取网页
from openpyxl import Workbook
headers = {
        'user-agent': "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36",
    }


def error521(url,data):
    response = requests.post(url,headers=headers,data=data)
    #获取__jsluid_s
    cookie_id=response.cookies
    cookie_id = '; '.join(['='.join(item) for item in cookie_id.items()])
    #首次页面提取js代码
    js_html=response.text
    js_func = ''.join(re.findall(r'<script>(.*?)</script>', js_html))#join将script全部连接起来
    #修改代码使得执行结果可得
    js_func=js_func.replace('eval(','var result=(')
    ctx= execjs.compile(js_func)#执行
    result=str(ctx.eval('result'))#查看结果
    #对结果的js代码进行剪切再次执行
    c_start=result.find('cookie')
    c_end=result.find("Path=/;'")+len("Path=/;'")
    cookie_js='var window={};var '+result[c_start:c_end]
    ctx=execjs.compile(cookie_js)
    cookie_js=ctx.eval('cookie')
    #得到__jsl_clearance
    cookie_clearance= cookie_js.split(';')[0]
    #拼接cookies
    cookies = {cookie_clearance.split('=')[0]: cookie_clearance.split('=')[1],cookie_id.split('=')[0]:cookie_id.split('=')[1]}
    return cookies


def dataProcess(ws,content,row):
    html=content.decode('utf-8')
    #文档里存在大量换行符、缩进
    #使用[\s\S]*?匹配任意文本。 “([\d\D]*)”、“([\w\W]*)”也可
    text=re.compile(r'(?<=<div id="flawList">)[\s\S]*?(?=<div class="pages clearfix">)').findall(html)[0]
    numbers=re.compile(r'(?<=/flaw/show/).*?(?=")').findall(text)#前20条
    titles=re.compile(r'(?<=title=").*?(?=">)').findall(text)
    times=re.compile(r'(?<=width="13%">)[\s\S]*?(?=</td>)').findall(text)#内含换行、tab。\n,\t
    ranks=re.compile(r'(?<=<span class=").*?(?="></span>)').findall(text)
    #类型判断是人工吗？？？？
    #types=re.compile(r'(?<=<span class=").*?(?="></span>)').findall(text)
    if len(numbers)==0:
        return 0
    for cnvd,title,time,rank in zip(numbers,titles,times,ranks):
        ws.cell(row, 1).value = row - 1
        ws.cell(row, 2).value = cnvd
        #好像打不开超链接
        #ws.cell(row,2).hyperlink = ('https://www.cnvd.org.cn/flaw/show/'+cnvd)#超链接
        ws.cell(row, 3).value = title
        ws.cell(row, 4).value = str(time).replace('\n','').replace('\t','')
        if rank=='red':
            ws.cell(row, 5).value = '高'
        if rank=='yellow':
            ws.cell(row, 5).value = '中'
        if rank=='green':
            ws.cell(row, 5).value = '低'
        types=re.compile('(?<=存在).*?漏洞').findall(title)
        if len(types)>0:
            type=''
            for t in types:
                type=type+t
            ws.cell(row, 6).value =type
        row+=1
    return 1

# 主函数
if __name__ == '__main__':
    #修改关键字
    key='firewall'
    file_name = key + ".xlsx"
    offset=0# offset=0开始
    max=50
    url = 'https://www.cnvd.org.cn/flaw/list.htm?flag=true'
    cookies=''
    response = ''
    flag=1
    row = 2
    #新技能！！openpyxl可以传参保存文件
    wb = Workbook()
    ws = wb.active
    ws.append(["序号", "cnvd编号", "名称", "时间", "等级", "漏洞类型"])
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['F'].width = 50
    print("******************start********************")
    while flag:
        data = {'keyword': key, 'max': max, 'offset': offset}
        try:
            cookies=error521(url,data)
            response = requests.post(url, headers=headers, cookies=cookies, data=data)
            flag = dataProcess(ws, response.content,row)
            offset+=max
            print('flag:' + str(flag) + '\toffset:' + str(offset))
            row+=max
        except Exception as e:
            print(e)
            traceback.print_exc()
            time.sleep(5)

    wb.save(file_name)
    print ("******************end********************")







