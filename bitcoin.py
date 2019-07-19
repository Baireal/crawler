# -*-coding:utf-8 -*-
import requests,re
from lxml import etree
from openpyxl import Workbook
headers = {
        'user-agent': "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36",
        'connection': "close",
    }

def btc(start,end,next_hash):
    #文件
    wb = Workbook()
    ws = wb.active
    ws.append(["block","hash","time","success", "input", "output", "amount"])
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 16
    #谁最多谁是row。。
    row = 2
    count=1
    #for i in range(start, end + 1):
    #避免异常出现写入错误，访问不成功继续请求
    while count<=end-start+1:
        #print("block:")
        try:
            url ='https://www.blockchain.com/btc/block/'+next_hash
            context=requests.get(url,headers=headers)#, timeout=60
            selector = etree.HTML(context.content)
            block_num=selector.xpath(u'//body[@class="opaque-nav"]/div/h1/text()')[0]
            block_num=str(block_num).replace('Block #','')
            #输出则表示该块请求成功
            print(block_num)
            #下一个块的哈希值
            block_hash = selector.xpath(u'//body[@class="opaque-nav"]/div/div[1]/div[2]/table/tr[4]/td[2]/a/text()')
            '''此处修改只访问一个块'''
            #更新下一个块哈希
            next_hash =str(block_hash[0]) #block_hash
            #block_time = selector.xpath(u'//body[@class="opaque-nav"]/div/div[1]/div[1]/table/tr[8]/td[2]/text()')
            trans=selector.xpath(u'//*[@class="txdiv"]')
            #按每笔交易选出list
            for tran in trans:
                hash = tran.xpath(u'./table/tr[1]/th/a/text()')
                time = tran.xpath(u'./table/tr[1]/th/span/text()')
                success = tran.xpath(u'./div/button/span/text()')

                ws.cell(row, 1).value = str(block_num)
                ws.cell(row, 2).value = str(hash[0])
                ws.cell(row, 3).value = str(time[0])
                ws.cell(row, 4).value = str(success[0])

                in_row=row
                out_row=row
                amount_row=row
                #merge_row = row #合并参数

                ints1 = tran.xpath(u'./table/tr[2]/td[1]/a/text()')
                for input in ints1:
                    ws.cell(row=in_row, column=5, value=str(input))
                    in_row = in_row + 1
                # 地址bc1
                ints2 = tran.xpath(u'./table/tr[2]/td[1]/text()')
                for input in ints2:
                    ws.cell(row=in_row, column=5, value=str(input))
                    in_row = in_row + 1
                #因为标签顺序错乱，最后选择正则匹配的顺序
                outs = tran.xpath(u'./table/tr[2]/td[3]')[0].xpath('string(.)')#获取全部td[3]下的text()
                outlist = re.compile(r'\bUnable\b|\b1\w{25,}\b|\b3\w{25,}\b|\bbc1\w*\b|(?<=BTC)Unable\b|(?<=BTC)1\w{25,}\b|(?<=BTC)3\w{25,}\b|(?<=BTC)bc1\w*\b').findall(outs)
                for output in outlist:
                    ws.cell(row=out_row, column=6, value=str(output))
                    out_row = out_row + 1
                amounts = tran.xpath(u'./table/tr[2]/td[3]/span/span/text()')
                for amount in amounts:
                    ws.cell(row=amount_row, column=7, value=str(amount))
                    amount_row = amount_row + 1

                row=max(in_row,out_row)

                #输出采用一对一
                '''outs1 = tran.xpath(u'./table/tr[2]/td[3]/a/text()')
                for output in outs1:
                    ws.cell(row=out_row, column=6, value=str(output))
                    out_row=out_row+1
                # 地址bc1
                outs2 = tran.xpath(u'./table/tr[2]/td[3]/text()')
                for output in outs2:
                    ws.cell(row=out_row, column=6, value=str(output))
                    out_row=out_row+1
                

                #amounts = tran.xpath(u'./table/tr[2]/td[3]/a/following-sibling::span/span/text()')
                amounts = tran.xpath(u'./table/tr[2]/td[3]/span/span')
                for amount in amounts:
                    am=amount.xpath(u'./text()')
                    output=amount.xpath(u'./../preceding-sibling::a[1]/text()')
                    if len(output)<1:
                        output=amount.xpath(u'./../preceding-sibling::br[1]/text()')
                    ws.cell(row=row, column=7, value=str(am[0]))
                    ws.cell(row=row,column=6,value=str(output[0]))
                    row = row + 1
                '''
                '''
                # 合并 block,time,hash,sum,input 很慢。。。。。
                ws.merge_cells(start_row=merge_row, start_column=1, end_row=row - 1, end_column=1)
                ws.merge_cells(start_row=merge_row, start_column=2, end_row=row - 1, end_column=2)
                ws.merge_cells(start_row=merge_row, start_column=3, end_row=row - 1, end_column=3)
                ws.merge_cells(start_row=merge_row, start_column=4, end_row=row - 1, end_column=4)
                
                #只可写入合并区域 左上 单元格
                ws.cell(merge_row, 1).value = str(i)
                ws.cell(merge_row, 2).value = str(block_time[0])
                ws.cell(merge_row, 3).value = str(hash[0])
                ws.cell(merge_row, 4).value = str(success[0])
                '''
        except Exception as e:
            print(e)
            continue#失败重来
        count=count+1

    wb.save("%d-%d.xlsx"%(start,end))
    return next_hash

# 主函数
if __name__ == '__main__':
    start = 584309#583189
    end = 584342
    step = 80
    print("**************** start *********************")
    #583189 hash
    hash="00000000000000000022c2315446c827c0445ed975ae2e4bc8e35a55d88ccfd7"#"000000000000000000010006e9fe5b90dacd169c686c4cbd6209e863a96eccaa"
    for i in range(start,end,step):
        if i+step<=end:
            print('start:%d end:%d' % (i,i+step-1))
            hash=btc(i,i+step-1,hash)
        else:
            print('start:%d end:%d' % (i, end))
            hash=btc(i,end,hash)
    print("****************  end  *********************")




